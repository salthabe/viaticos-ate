from fastapi import FastAPI, HTTPException, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel
from typing import Optional, List
import os
from datetime import datetime
import calendar
from contextlib import contextmanager
import tempfile

app = FastAPI(title="Sistema de Viáticos ATE")

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
EXPORTS_DIR = os.path.join(tempfile.gettempdir(), "viaticos_exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

static_dir = os.path.join(BASE_DIR, "static")
if os.path.exists(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")

# ── DATABASE: PostgreSQL en Railway, SQLite en local ─────────────────────────

DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

USE_PG = bool(DATABASE_URL)
PH     = "%s" if USE_PG else "?"

if USE_PG:
    import psycopg2, psycopg2.extras

    @contextmanager
    def get_db():
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = False
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _rows(cur):
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]

    def _row(cur):
        if not cur.description:
            return None
        cols = [d[0] for d in cur.description]
        r = cur.fetchone()
        return dict(zip(cols, r)) if r else None

    def _insert(conn, sql, params):
        cur = conn.cursor()
        cur.execute(sql + " RETURNING id", params)
        return cur.fetchone()[0]

    def _year_filter(col):  return f"EXTRACT(YEAR  FROM {col}) = {PH}"
    def _month_filter(col): return f"EXTRACT(MONTH FROM {col}) = {PH}"
    def _yp(v): return v          # año/mes como int
    def _mp(v): return v

    def init_db():
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS agentes (
                    id SERIAL PRIMARY KEY, nombre TEXT NOT NULL,
                    cuit TEXT, cbu TEXT, banco TEXT, alias TEXT,
                    tope_mensual REAL DEFAULT 0, activo INTEGER DEFAULT 1,
                    creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS categorias (
                    id SERIAL PRIMARY KEY, nombre TEXT NOT NULL UNIQUE, activa INTEGER DEFAULT 1
                );
                CREATE TABLE IF NOT EXISTS tickets (
                    id SERIAL PRIMARY KEY,
                    agente_id INTEGER NOT NULL REFERENCES agentes(id),
                    fecha_gasto DATE NOT NULL, categoria_id INTEGER REFERENCES categorias(id),
                    comprobante TEXT, descripcion TEXT, valor REAL NOT NULL,
                    estado TEXT DEFAULT 'pendiente', motivo_rechazo TEXT, valor_aprobado REAL,
                    periodo_pago_id INTEGER REFERENCES periodos_pago(id),
                    creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    revisado_en TIMESTAMP, revisado_por TEXT
                );
                CREATE TABLE IF NOT EXISTS periodos (
                    id SERIAL PRIMARY KEY,
                    agente_id INTEGER NOT NULL REFERENCES agentes(id),
                    anio INTEGER NOT NULL, mes INTEGER NOT NULL,
                    tope_override REAL, cerrado INTEGER DEFAULT 0,
                    UNIQUE(agente_id, anio, mes)
                );
                CREATE TABLE IF NOT EXISTS adjuntos (
                    id SERIAL PRIMARY KEY,
                    ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    nombre_archivo TEXT NOT NULL,
                    mime_type TEXT,
                    datos BYTEA NOT NULL,
                    subido_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS historial (
                    id SERIAL PRIMARY KEY,
                    ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    campo TEXT NOT NULL,
                    valor_anterior TEXT,
                    valor_nuevo TEXT,
                    usuario TEXT,
                    fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS periodos_cierre (
                    id SERIAL PRIMARY KEY,
                    anio INTEGER NOT NULL,
                    mes INTEGER NOT NULL,
                    cerrado INTEGER DEFAULT 0,
                    cerrado_por TEXT,
                    cerrado_en TIMESTAMP,
                    UNIQUE(anio, mes)
                );
                CREATE TABLE IF NOT EXISTS periodos_pago (
                    id SERIAL PRIMARY KEY,
                    numero INTEGER NOT NULL,
                    nombre TEXT NOT NULL,
                    descripcion TEXT,
                    estado TEXT DEFAULT 'abierto',
                    creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    cerrado_en TIMESTAMP,
                    cerrado_por TEXT
                );
                CREATE TABLE IF NOT EXISTS actas (
                    id SERIAL PRIMARY KEY,
                    numero_acta TEXT NOT NULL,
                    fecha DATE NOT NULL,
                    tipo TEXT NOT NULL,
                    titulo TEXT NOT NULL,
                    cuerpo TEXT,
                    participantes TEXT,
                    redactado_por TEXT,
                    creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    modificado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cur.execute("""
                INSERT INTO categorias (nombre)
                VALUES ('Transporte'),('Alojamiento'),('Alimentación'),
                       ('Combustible'),('Peajes'),('Materiales'),('Otro')
                ON CONFLICT (nombre) DO NOTHING;
            """)

else:
    import sqlite3
    DB_PATH = os.path.join(BASE_DIR, "viaticos.db")

    @contextmanager
    def get_db():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _rows(cur): return [dict(r) for r in cur.fetchall()]
    def _row(cur):
        r = cur.fetchone()
        return dict(r) if r else None

    def _insert(conn, sql, params):
        cur = conn.cursor()
        cur.execute(sql, params)
        return cur.lastrowid

    def _year_filter(col):  return f"strftime('%Y', {col}) = {PH}"
    def _month_filter(col): return f"strftime('%m', {col}) = {PH}"
    def _yp(v): return str(v)
    def _mp(v): return f"{v:02d}"

    def init_db():
        with get_db() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS agentes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL,
                    cuit TEXT, cbu TEXT, banco TEXT, alias TEXT,
                    tope_mensual REAL DEFAULT 0, activo INTEGER DEFAULT 1,
                    creado_en TEXT DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS categorias (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre TEXT NOT NULL UNIQUE, activa INTEGER DEFAULT 1
                );
                CREATE TABLE IF NOT EXISTS tickets (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    agente_id INTEGER NOT NULL REFERENCES agentes(id),
                    fecha_gasto TEXT NOT NULL, categoria_id INTEGER REFERENCES categorias(id),
                    comprobante TEXT, descripcion TEXT, valor REAL NOT NULL,
                    estado TEXT DEFAULT 'pendiente', motivo_rechazo TEXT, valor_aprobado REAL,
                    periodo_pago_id INTEGER REFERENCES periodos_pago(id),
                    creado_en TEXT DEFAULT CURRENT_TIMESTAMP, revisado_en TEXT, revisado_por TEXT
                );
                CREATE TABLE IF NOT EXISTS periodos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    agente_id INTEGER NOT NULL REFERENCES agentes(id),
                    anio INTEGER NOT NULL, mes INTEGER NOT NULL,
                    tope_override REAL, cerrado INTEGER DEFAULT 0,
                    UNIQUE(agente_id, anio, mes)
                );
                CREATE TABLE IF NOT EXISTS adjuntos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    nombre_archivo TEXT NOT NULL,
                    mime_type TEXT,
                    datos BLOB NOT NULL,
                    subido_en TEXT DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS historial (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    campo TEXT NOT NULL,
                    valor_anterior TEXT,
                    valor_nuevo TEXT,
                    usuario TEXT,
                    fecha TEXT DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS periodos_cierre (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    anio INTEGER NOT NULL,
                    mes INTEGER NOT NULL,
                    cerrado INTEGER DEFAULT 0,
                    cerrado_por TEXT,
                    cerrado_en TEXT,
                    UNIQUE(anio, mes)
                );
                CREATE TABLE IF NOT EXISTS periodos_pago (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero INTEGER NOT NULL,
                    nombre TEXT NOT NULL,
                    descripcion TEXT,
                    estado TEXT DEFAULT 'abierto',
                    creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
                    cerrado_en TEXT,
                    cerrado_por TEXT
                );
                CREATE TABLE IF NOT EXISTS actas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero_acta TEXT NOT NULL,
                    fecha TEXT NOT NULL,
                    tipo TEXT NOT NULL,
                    titulo TEXT NOT NULL,
                    cuerpo TEXT,
                    participantes TEXT,
                    redactado_por TEXT,
                    creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
                    modificado_en TEXT DEFAULT CURRENT_TIMESTAMP
                );
                INSERT OR IGNORE INTO categorias (nombre) VALUES
                    ('Transporte'),('Alojamiento'),('Alimentación'),
                    ('Combustible'),('Peajes'),('Materiales'),('Otro');
            """)

init_db()

# ── MODELS ────────────────────────────────────────────────────────────────────

class AgenteCreate(BaseModel):
    nombre: str
    cuit: Optional[str] = None
    cbu: Optional[str] = None
    banco: Optional[str] = None
    alias: Optional[str] = None
    tope_mensual: float = 0

class AgenteUpdate(BaseModel):
    nombre: Optional[str] = None
    cuit: Optional[str] = None
    cbu: Optional[str] = None
    banco: Optional[str] = None
    alias: Optional[str] = None
    tope_mensual: Optional[float] = None
    activo: Optional[int] = None

class TicketCreate(BaseModel):
    agente_id: int
    fecha_gasto: str
    categoria_id: Optional[int] = None
    comprobante: Optional[str] = None
    descripcion: Optional[str] = None
    valor: float

class TicketRevision(BaseModel):
    estado: str
    motivo_rechazo: Optional[str] = None
    valor_aprobado: Optional[float] = None
    revisado_por: Optional[str] = None

class PeriodoTope(BaseModel):
    agente_id: int
    anio: int
    mes: int
    tope_override: Optional[float] = None

# ── AGENTES ───────────────────────────────────────────────────────────────────

@app.get("/api/agentes")
def listar_agentes(solo_activos: bool = True):
    with get_db() as conn:
        cur = conn.cursor()
        q = "SELECT * FROM agentes" + (" WHERE activo=1" if solo_activos else "") + " ORDER BY nombre"
        cur.execute(q)
        return _rows(cur)

@app.post("/api/agentes")
def crear_agente(data: AgenteCreate):
    with get_db() as conn:
        new_id = _insert(conn,
            f"INSERT INTO agentes (nombre,cuit,cbu,banco,alias,tope_mensual) VALUES ({PH},{PH},{PH},{PH},{PH},{PH})",
            (data.nombre, data.cuit, data.cbu, data.banco, data.alias, data.tope_mensual))
        return {"id": new_id, "mensaje": "Agente creado"}

@app.put("/api/agentes/{agente_id}")
def actualizar_agente(agente_id: int, data: AgenteUpdate):
    campos = {k: v for k, v in data.dict().items() if v is not None}
    if not campos:
        raise HTTPException(400, "Sin datos para actualizar")
    with get_db() as conn:
        cur = conn.cursor()
        sets = ", ".join(f"{k}={PH}" for k in campos)
        cur.execute(f"UPDATE agentes SET {sets} WHERE id={PH}", (*campos.values(), agente_id))
    return {"mensaje": "Actualizado"}

@app.get("/api/agentes/{agente_id}")
def obtener_agente(agente_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM agentes WHERE id={PH}", (agente_id,))
        row = _row(cur)
        if not row:
            raise HTTPException(404, "Agente no encontrado")
        return row

# ── TICKETS ───────────────────────────────────────────────────────────────────

@app.get("/api/tickets")
def listar_tickets(agente_id: Optional[int] = None, estado: Optional[str] = None,
                   anio: Optional[int] = None, mes: Optional[int] = None,
                   periodo_pago_id: Optional[int] = None):
    with get_db() as conn:
        cur = conn.cursor()
        q = """SELECT t.*, a.nombre as agente_nombre, c.nombre as categoria_nombre,
                      pp.nombre as periodo_pago_nombre, pp.numero as periodo_pago_numero
               FROM tickets t
               LEFT JOIN agentes a ON t.agente_id = a.id
               LEFT JOIN categorias c ON t.categoria_id = c.id
               LEFT JOIN periodos_pago pp ON t.periodo_pago_id = pp.id
               WHERE 1=1"""
        params = []
        if agente_id:
            q += f" AND t.agente_id={PH}"; params.append(agente_id)
        if estado:
            q += f" AND t.estado={PH}"; params.append(estado)
        if anio:
            q += f" AND {_year_filter('t.fecha_gasto')}"; params.append(_yp(anio))
        if mes:
            q += f" AND {_month_filter('t.fecha_gasto')}"; params.append(_mp(mes))
        if periodo_pago_id:
            q += f" AND t.periodo_pago_id={PH}"; params.append(periodo_pago_id)
        q += " ORDER BY t.fecha_gasto DESC"
        cur.execute(q, params)
        rows = _rows(cur)
        for r in rows:
            if r.get("fecha_gasto") and not isinstance(r["fecha_gasto"], str):
                r["fecha_gasto"] = r["fecha_gasto"].strftime("%Y-%m-%d")
        return rows

@app.post("/api/tickets")
def crear_ticket(data: TicketCreate):
    with get_db() as conn:
        # Auto-asignar al período de pago abierto
        pp = _get_periodo_abierto(conn)
        pp_id = pp["id"] if pp else None
        new_id = _insert(conn,
            f"INSERT INTO tickets (agente_id,fecha_gasto,categoria_id,comprobante,descripcion,valor,periodo_pago_id) VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH})",
            (data.agente_id, data.fecha_gasto, data.categoria_id, data.comprobante, data.descripcion, data.valor, pp_id))
        return {"id": new_id, "mensaje": "Ticket creado", "periodo_pago_id": pp_id}

@app.put("/api/tickets/{ticket_id}/revision")
def revisar_ticket(ticket_id: int, data: TicketRevision):
    if data.estado not in ("aprobado", "rechazado", "debito_parcial", "pagado"):
        raise HTTPException(400, "Estado inválido")
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM tickets WHERE id={PH}", (ticket_id,))
        ticket = _row(cur)
        if not ticket:
            raise HTTPException(404, "Ticket no encontrado")
        valor_aprobado = (data.valor_aprobado if data.estado == "debito_parcial"
                          else ticket["valor"] if data.estado == "aprobado"
                          else ticket.get("valor_aprobado") if data.estado == "pagado"
                          else 0)
        cur.execute(f"""UPDATE tickets SET estado={PH}, motivo_rechazo={PH}, valor_aprobado={PH},
                    revisado_en=CURRENT_TIMESTAMP, revisado_por={PH} WHERE id={PH}""",
                    (data.estado, data.motivo_rechazo, valor_aprobado, data.revisado_por, ticket_id))
        # Log history
        if ticket["estado"] != data.estado:
            _registrar_historial(conn, ticket_id, "estado", ticket["estado"], data.estado, data.revisado_por)
        if data.estado == "debito_parcial" and data.valor_aprobado is not None:
            _registrar_historial(conn, ticket_id, "valor_aprobado", ticket.get("valor_aprobado"), valor_aprobado, data.revisado_por)
        if data.motivo_rechazo:
            _registrar_historial(conn, ticket_id, "motivo_rechazo", ticket.get("motivo_rechazo"), data.motivo_rechazo, data.revisado_por)
    return {"mensaje": "Ticket revisado"}

@app.delete("/api/tickets/{ticket_id}")
def eliminar_ticket(ticket_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM tickets WHERE id={PH}", (ticket_id,))
    return {"mensaje": "Ticket eliminado"}

# ── CATEGORÍAS ────────────────────────────────────────────────────────────────

@app.get("/api/categorias")
def listar_categorias():
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM categorias WHERE activa=1 ORDER BY nombre")
        return _rows(cur)

# ── RESUMEN ───────────────────────────────────────────────────────────────────

@app.get("/api/resumen")
def resumen(anio: int, mes: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT a.id, a.nombre, a.tope_mensual, a.cbu, a.banco, a.alias, a.cuit,
                COALESCE(SUM(CASE WHEN t.estado='pendiente'      THEN t.valor          END), 0) as total_pendiente,
                COALESCE(SUM(CASE WHEN t.estado='aprobado'       THEN t.valor_aprobado END), 0) as total_aprobado,
                COALESCE(SUM(CASE WHEN t.estado='debito_parcial' THEN t.valor_aprobado END), 0) as total_debito,
                COALESCE(SUM(CASE WHEN t.estado='pagado'         THEN t.valor_aprobado END), 0) as total_pagado,
                COALESCE(SUM(CASE WHEN t.estado='rechazado'      THEN t.valor          END), 0) as total_rechazado,
                COUNT(CASE WHEN t.estado='pendiente'    THEN 1 END) as cant_pendientes,
                COUNT(CASE WHEN t.estado='aprobado'     THEN 1 END) as cant_aprobados,
                COUNT(CASE WHEN t.estado='debito_parcial' THEN 1 END) as cant_debito,
                COUNT(CASE WHEN t.estado='pagado'       THEN 1 END) as cant_pagados,
                COUNT(CASE WHEN t.estado='rechazado'    THEN 1 END) as cant_rechazados
            FROM agentes a
            LEFT JOIN tickets t ON a.id = t.agente_id
                AND {_year_filter('t.fecha_gasto')}
                AND {_month_filter('t.fecha_gasto')}
            WHERE a.activo = 1
            GROUP BY a.id, a.nombre, a.tope_mensual, a.cbu, a.banco, a.alias, a.cuit
            ORDER BY a.nombre
        """, (_yp(anio), _mp(mes)))
        rows = _rows(cur)

        result = []
        for d in rows:
            cur.execute(f"SELECT tope_override FROM periodos WHERE agente_id={PH} AND anio={PH} AND mes={PH}",
                        (d["id"], anio, mes))
            ov = _row(cur)
            tope = d["tope_mensual"]
            if ov and ov.get("tope_override") is not None:
                tope = ov["tope_override"]
            subtotal = d["total_aprobado"] + d["total_debito"] + d["total_pagado"]
            d["tope_efectivo"] = tope
            d["a_transferir"]  = min(subtotal, tope) if tope > 0 else subtotal
            d["excedente"]     = max(0, subtotal - tope) if tope > 0 else 0
            result.append(d)
        return result

@app.get("/api/resumen/semanal")
def resumen_semanal(anio: int, mes: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT t.*, a.nombre as agente_nombre
            FROM tickets t JOIN agentes a ON t.agente_id = a.id
            WHERE {_year_filter('t.fecha_gasto')} AND {_month_filter('t.fecha_gasto')}
              AND t.estado IN ('aprobado','debito_parcial','pagado')
            ORDER BY t.fecha_gasto
        """, (_yp(anio), _mp(mes)))
        tickets = _rows(cur)

    semanas = {1: [], 2: [], 3: [], 4: [], 5: []}
    for t in tickets:
        fg = t["fecha_gasto"]
        if not isinstance(fg, str):
            fg = fg.strftime("%Y-%m-%d")
            t["fecha_gasto"] = fg
        day    = int(fg.split("-")[2])
        semana = min((day - 1) // 7 + 1, 5)
        semanas[semana].append(t)

    return [{"semana": s, "tickets": items, "total": sum(i.get("valor_aprobado") or 0 for i in items)}
            for s, items in semanas.items() if items]

# ── TOPES ─────────────────────────────────────────────────────────────────────

@app.post("/api/topes")
def configurar_tope(data: PeriodoTope):
    with get_db() as conn:
        cur = conn.cursor()
        if USE_PG:
            cur.execute("""
                INSERT INTO periodos (agente_id, anio, mes, tope_override) VALUES (%s,%s,%s,%s)
                ON CONFLICT (agente_id, anio, mes) DO UPDATE SET tope_override = EXCLUDED.tope_override
            """, (data.agente_id, data.anio, data.mes, data.tope_override))
        else:
            cur.execute("""
                INSERT INTO periodos (agente_id, anio, mes, tope_override) VALUES (?,?,?,?)
                ON CONFLICT(agente_id,anio,mes) DO UPDATE SET tope_override=excluded.tope_override
            """, (data.agente_id, data.anio, data.mes, data.tope_override))
    return {"mensaje": "Tope configurado"}


# ── ELIMINAR AGENTE ───────────────────────────────────────────────────────────

@app.delete("/api/agentes/{agente_id}")
def eliminar_agente(agente_id: int, forzar: bool = False):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT COUNT(*) as n FROM tickets WHERE agente_id={PH}", (agente_id,))
        row = _row(cur)
        cant = row["n"] if row else 0
        if cant > 0 and not forzar:
            raise HTTPException(409, f"El agente tiene {cant} ticket(s). Usar forzar=true para eliminar junto con sus tickets.")
        if forzar:
            cur.execute(f"DELETE FROM periodos WHERE agente_id={PH}", (agente_id,))
            cur.execute(f"DELETE FROM tickets WHERE agente_id={PH}", (agente_id,))
        cur.execute(f"DELETE FROM agentes WHERE id={PH}", (agente_id,))
    return {"mensaje": "Agente eliminado"}


# ── IMPORTACIÓN MASIVA EXCEL ──────────────────────────────────────────────────

class ImportResult(BaseModel):
    insertados: int
    duplicados: int
    errores: list
    detalle: list

@app.get("/api/importar/modelo")
def descargar_modelo():
    """Genera y descarga el Excel modelo para carga masiva de tickets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "SINDICATO ATE — Plantilla de Importación Masiva de Tickets"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["agente_nombre *", "fecha_gasto * (DD/MM/AAAA)", "categoria", "comprobante", "descripcion", "valor *", "notas_internas"]
    col_widths = [28, 18, 16, 18, 32, 14, 24]
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2d6a9f")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[chr(64+c)].width = w
    ws.row_dimensions[2].height = 32

    # Instrucciones en fila 3
    ws.merge_cells("A3:G3")
    ws["A3"] = "⚠ IMPORTANTE: agente_nombre debe coincidir EXACTAMENTE con el nombre registrado en el sistema. fecha_gasto formato: 15/03/2026 (DD/MM/AAAA). valor: solo números."
    ws["A3"].font = Font(size=10, color="7B341E")
    ws["A3"].fill = PatternFill("solid", fgColor="FEF3C7")
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 28

    # Filas de ejemplo
    ejemplos = [
        ("Seccional Buenos Aires", "05/03/2026", "Transporte", "FC-0001-00001", "Viaje a sede central", 12500, ""),
        ("Seccional La Plata",     "07/03/2026", "Combustible","FC-0004-00002", "Carga nafta camioneta", 8500, ""),
        ("Seccional Rosario",      "11/03/2026", "Alojamiento","FC-0002-00015", "2 noches congreso", 18000, ""),
    ]
    for ri, ej in enumerate(ejemplos, 4):
        for c, v in enumerate(ej, 1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor="EBF3FB")
            if c == 6:
                cell.number_format = '#,##0.00'

    # Hoja de referencia de agentes
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT nombre FROM agentes WHERE activo=1 ORDER BY nombre")
        agentes_lista = [r["nombre"] for r in _rows(cur)]
        cur.execute("SELECT nombre FROM categorias WHERE activa=1 ORDER BY nombre")
        cats_lista = [r["nombre"] for r in _rows(cur)]

    ws2 = wb.create_sheet("Referencia")
    ws2["A1"] = "Agentes registrados (copiar nombre exacto)"; ws2["A1"].font = Font(bold=True, color="FFFFFF"); ws2["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws2["B1"] = "Categorías disponibles"; ws2["B1"].font = Font(bold=True, color="FFFFFF"); ws2["B1"].fill = PatternFill("solid", fgColor="1a3a5c")
    for i, nombre in enumerate(agentes_lista, 2):
        ws2.cell(row=i, column=1, value=nombre)
    for i, cat in enumerate(cats_lista, 2):
        ws2.cell(row=i, column=2, value=cat)
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 18

    filename = "modelo_importacion_tickets_ATE.xlsx"
    path = os.path.join(EXPORTS_DIR, filename)
    wb.save(path)
    return FileResponse(path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/importar/tickets")
async def importar_tickets(request: dict):
    """
    Recibe lista de tickets desde el frontend (parseado del Excel).
    Verifica duplicados por (agente_nombre, fecha_gasto, comprobante, valor).
    """
    filas = request.get("filas", [])
    if not filas:
        raise HTTPException(400, "No hay filas para importar")

    insertados = 0
    duplicados = 0
    errores = []
    detalle = []

    with get_db() as conn:
        cur = conn.cursor()

        # Obtener período activo
        pp = _get_periodo_abierto(conn)
        pp_id = pp["id"] if pp else None

        # Cargar mapa agente_nombre -> id
        cur.execute("SELECT id, nombre FROM agentes WHERE activo=1")
        agentes_map = {r["nombre"].strip().lower(): r["id"] for r in _rows(cur)}

        # Cargar mapa categoria_nombre -> id
        cur.execute("SELECT id, nombre FROM categorias WHERE activa=1")
        cats_map = {r["nombre"].strip().lower(): r["id"] for r in _rows(cur)}

        for i, fila in enumerate(filas, 1):
            fila_num = i + 3  # filas 4+ en el Excel (1=header titulo, 2=headers, 3=instruccion)

            agente_nombre = str(fila.get("agente_nombre", "")).strip()
            fecha_gasto   = str(fila.get("fecha_gasto", "")).strip()
            categoria     = str(fila.get("categoria", "")).strip()
            comprobante   = str(fila.get("comprobante", "")).strip() or None
            descripcion   = str(fila.get("descripcion", "")).strip() or None
            valor_raw     = fila.get("valor", "")

            # Validaciones básicas
            if not agente_nombre:
                errores.append(f"Fila {fila_num}: agente_nombre vacío"); continue
            if not fecha_gasto:
                errores.append(f"Fila {fila_num}: fecha_gasto vacía"); continue

            # Buscar agente (case-insensitive)
            agente_id = agentes_map.get(agente_nombre.lower())
            if not agente_id:
                errores.append(f"Fila {fila_num}: Agente '{agente_nombre}' no encontrado en el sistema"); continue

            # Validar fecha — aceptar DD/MM/AAAA (principal) o AAAA-MM-DD (fallback)
            try:
                dt = datetime.strptime(fecha_gasto, "%d/%m/%Y")
                fecha_gasto = dt.strftime("%Y-%m-%d")
            except ValueError:
                try:
                    datetime.strptime(fecha_gasto, "%Y-%m-%d")  # ya está en formato correcto
                except:
                    errores.append(f"Fila {fila_num}: Fecha inválida '{fecha_gasto}' — usar formato DD/MM/AAAA"); continue

            # Validar valor
            try:
                valor = float(str(valor_raw).replace(",", ".").replace("$", "").strip())
                if valor <= 0:
                    raise ValueError()
            except:
                errores.append(f"Fila {fila_num}: Valor inválido '{valor_raw}'"); continue

            # Categoría opcional
            categoria_id = cats_map.get(categoria.lower()) if categoria else None

            # Verificar duplicado: mismo agente + fecha + comprobante + valor
            dup_check = [agente_id, fecha_gasto, valor]
            dup_q = f"SELECT id FROM tickets WHERE agente_id={PH} AND fecha_gasto={PH} AND valor={PH}"
            if comprobante:
                dup_q += f" AND comprobante={PH}"
                dup_check.append(comprobante)
            cur.execute(dup_q, dup_check)
            existing = _row(cur)
            if existing:
                duplicados += 1
                detalle.append({"fila": fila_num, "estado": "duplicado", "agente": agente_nombre,
                                 "fecha": fecha_gasto, "valor": valor, "ticket_existente": existing["id"]})
                continue

            # Insertar
            new_id = _insert(conn,
                f"INSERT INTO tickets (agente_id,fecha_gasto,categoria_id,comprobante,descripcion,valor,periodo_pago_id) VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH})",
                (agente_id, fecha_gasto, categoria_id, comprobante, descripcion, valor, pp_id))
            insertados += 1
            detalle.append({"fila": fila_num, "estado": "insertado", "id": new_id,
                            "agente": agente_nombre, "fecha": fecha_gasto, "valor": valor})

    return {"insertados": insertados, "duplicados": duplicados,
            "errores": errores, "detalle": detalle}


# ── EXPORTACIÓN EXCEL PERSONALIZADA ──────────────────────────────────────────

@app.get("/api/exportar/excel/custom")
def exportar_excel_custom(
    anio: Optional[int] = None,
    mes: Optional[int] = None,
    agente_id: Optional[int] = None,
    estado: Optional[List[str]] = Query(default=None),
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    incluir_transferencias: bool = True,
    incluir_tickets: bool = True,
    incluir_semanal: bool = False,
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Obtener tickets con filtros
    with get_db() as conn:
        cur = conn.cursor()
        q = """SELECT t.*, a.nombre as agente_nombre, a.cuit, a.cbu, a.banco, a.alias,
                      c.nombre as categoria_nombre
               FROM tickets t
               LEFT JOIN agentes a ON t.agente_id = a.id
               LEFT JOIN categorias c ON t.categoria_id = c.id
               WHERE 1=1"""
        params = []
        if agente_id:
            q += f" AND t.agente_id={PH}"; params.append(agente_id)
        if estado:
            # Normalizar: puede llegar como str o list dependiendo de FastAPI/querystring
            if isinstance(estado, str):
                estado = [estado]
            placeholders = ",".join([PH] * len(estado))
            q += f" AND t.estado IN ({placeholders})"; params.extend(estado)
        if anio:
            q += f" AND {_year_filter('t.fecha_gasto')}"; params.append(_yp(anio))
        if mes:
            q += f" AND {_month_filter('t.fecha_gasto')}"; params.append(_mp(mes))
        if desde:
            q += f" AND t.fecha_gasto >= {PH}"; params.append(desde)
        if hasta:
            q += f" AND t.fecha_gasto <= {PH}"; params.append(hasta)
        q += " ORDER BY a.nombre, t.fecha_gasto"
        cur.execute(q, params)
        tickets = _rows(cur)
        for t in tickets:
            if t.get("fecha_gasto") and not isinstance(t["fecha_gasto"], str):
                t["fecha_gasto"] = t["fecha_gasto"].strftime("%Y-%m-%d")

    # Título dinámico
    partes = []
    if anio and mes:  partes.append(f"{calendar.month_name[mes].upper()} {anio}")
    elif anio:        partes.append(str(anio))
    if estado:
        if isinstance(estado, str): estado = [estado]
        partes.append(", ".join(e.upper() for e in estado))
    if desde or hasta:
        partes.append(f"{desde or ''} al {hasta or ''}")
    titulo_filtro = " — ".join(partes) if partes else "TODOS LOS REGISTROS"

    wb = openpyxl.Workbook()
    first = True

    colores = {"aprobado":"D4EDDA","rechazado":"F8D7DA","debito_parcial":"FFF3CD","pendiente":"FFFFFF","pagado":"E0F2FE"}

    if incluir_tickets:
        ws = wb.active if first else wb.create_sheet("Tickets")
        ws.title = "Tickets"
        first = False

        ws.merge_cells("A1:I1")
        ws["A1"] = f"SINDICATO ATE — TICKETS — {titulo_filtro}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        hdrs = ["Agente","Fecha","Categoría","Comprobante","Descripción","Valor","Estado","Aprobado","Motivo"]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2d6a9f")
            cell.alignment = Alignment(horizontal="center")

        for ri, t in enumerate(tickets, 3):
            color = colores.get(t["estado"], "FFFFFF")
            vals = [t["agente_nombre"], t["fecha_gasto"], t.get("categoria_nombre",""),
                    t.get("comprobante",""), t.get("descripcion",""), t["valor"],
                    t["estado"].upper(), t.get("valor_aprobado") or "", t.get("motivo_rechazo","")]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=c, value=v)
                cell.fill = PatternFill("solid", fgColor=color)
                if c in (6,8): cell.number_format = '"$"#,##0.00'

        # Fila totales
        tr = len(tickets) + 3
        ws.cell(row=tr, column=5, value="TOTAL").font = Font(bold=True)
        ws.cell(row=tr, column=6, value=sum(t["valor"] for t in tickets)).number_format = '"$"#,##0.00'
        ws.cell(row=tr, column=6).font = Font(bold=True)
        ws.cell(row=tr, column=8, value=sum(t.get("valor_aprobado") or 0 for t in tickets)).number_format = '"$"#,##0.00'
        ws.cell(row=tr, column=8).font = Font(bold=True)

        for col, w in zip("ABCDEFGHI", [28,12,16,18,32,14,14,14,32]):
            ws.column_dimensions[col].width = w

    if incluir_transferencias and anio and mes:
        ws2 = wb.active if first else wb.create_sheet("Transferencias")
        ws2.title = "Transferencias"
        first = False

        res = resumen(anio, mes)
        if agente_id:
            res = [r for r in res if r["id"] == agente_id]

        ws2.merge_cells("A1:H1")
        ws2["A1"] = f"SINDICATO ATE — TRANSFERENCIAS — {titulo_filtro}"
        ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws2["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 26

        hdrs2 = ["Agente","CUIT","CBU","Banco/Alias","Total Presentado","Aprobado","Tope","A TRANSFERIR"]
        for c, h in enumerate(hdrs2, 1):
            cell = ws2.cell(row=2, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2d6a9f")

        total_t = 0
        for ri, ag in enumerate(res, 3):
            tp = ag["total_aprobado"] + ag["total_debito"] + ag["total_pendiente"]
            vals = [ag["nombre"], ag.get("cuit",""), ag.get("cbu",""),
                    ag.get("banco","") or ag.get("alias",""),
                    tp, ag["total_aprobado"]+ag["total_debito"], ag["tope_efectivo"], ag["a_transferir"]]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=ri, column=c, value=v)
                if c in (5,6,7,8): cell.number_format = '"$"#,##0.00'
                if ri % 2 == 0: cell.fill = PatternFill("solid", fgColor="EBF3FB")
            total_t += ag["a_transferir"]
        tr2 = len(res) + 3
        ws2.cell(row=tr2, column=1, value="TOTAL").font = Font(bold=True)
        ct = ws2.cell(row=tr2, column=8, value=total_t)
        ct.font = Font(bold=True, color="FFFFFF"); ct.fill = PatternFill("solid", fgColor="1a3a5c")
        ct.number_format = '"$"#,##0.00'
        for i, w in enumerate([30,16,26,20,16,16,14,16], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    if incluir_semanal and anio and mes:
        ws3 = wb.active if first else wb.create_sheet("Semanal")
        ws3.title = "Semanal"; first = False
        ws3.merge_cells("A1:D1")
        ws3["A1"] = f"RESUMEN SEMANAL — {titulo_filtro}"
        ws3["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws3["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
        ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
        row = 2
        for s in resumen_semanal(anio, mes):
            ws3.cell(row=row, column=1, value=f"Semana {s['semana']}").font = Font(bold=True, color="FFFFFF")
            ws3.cell(row=row, column=1).fill = PatternFill("solid", fgColor="2d6a9f")
            ws3.merge_cells(f"A{row}:D{row}"); row += 1
            for c, h in enumerate(["Agente","Fecha","Descripción","Aprobado"], 1):
                ws3.cell(row=row, column=c, value=h).font = Font(bold=True)
            row += 1
            for t in s["tickets"]:
                ws3.cell(row=row,column=1,value=t["agente_nombre"])
                ws3.cell(row=row,column=2,value=t["fecha_gasto"])
                ws3.cell(row=row,column=3,value=t.get("descripcion",""))
                ws3.cell(row=row,column=4,value=t.get("valor_aprobado",0)).number_format='"$"#,##0.00'
                row += 1
            ws3.cell(row=row,column=3,value="TOTAL").font=Font(bold=True)
            ws3.cell(row=row,column=4,value=s["total"]).font=Font(bold=True)
            ws3.cell(row=row,column=4).number_format='"$"#,##0.00'
            row += 2
        for col in "ABCD": ws3.column_dimensions[col].width = 28

    # Partes del nombre de archivo
    fname_parts = ["export_ATE"]
    if anio: fname_parts.append(str(anio))
    if mes:  fname_parts.append(f"{mes:02d}")
    if estado: fname_parts.append("_".join(estado) if isinstance(estado, list) else estado)
    filename = "_".join(fname_parts) + ".xlsx"
    path = os.path.join(EXPORTS_DIR, filename)
    wb.save(path)
    return FileResponse(path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── EXPORTACIÓN EXCEL ─────────────────────────────────────────────────────────

@app.get("/api/exportar/excel")
def exportar_excel(anio: int, mes: int):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    resumen_data = resumen(anio, mes)
    nombre_mes   = calendar.month_name[mes]
    wb = openpyxl.Workbook()

    # Hoja 1: Transferencias
    ws1 = wb.active; ws1.title = "Transferencias"
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"SINDICATO ATE — REINTEGRO DE VIÁTICOS {nombre_mes.upper()} {anio}"
    ws1["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    hdrs = ["Agente","CUIT","CBU","Banco / Alias","Total Presentado","Total Aprobado","Tope Mensual","A TRANSFERIR"]
    for c, h in enumerate(hdrs, 1):
        cell = ws1.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2d6a9f")
        cell.alignment = Alignment(horizontal="center")

    total_transf = 0
    for ri, ag in enumerate(resumen_data, 3):
        tp = ag["total_aprobado"] + ag["total_debito"] + ag["total_pendiente"]
        vals = [ag["nombre"], ag.get("cuit",""), ag.get("cbu",""),
                ag.get("banco","") or ag.get("alias",""),
                tp, ag["total_aprobado"]+ag["total_debito"], ag["tope_efectivo"], ag["a_transferir"]]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=c, value=v)
            if c in (5,6,7,8): cell.number_format = '"$"#,##0.00'
            if ri % 2 == 0:    cell.fill = PatternFill("solid", fgColor="EBF3FB")
        total_transf += ag["a_transferir"]

    tr = len(resumen_data) + 3
    ws1.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ct = ws1.cell(row=tr, column=8, value=total_transf)
    ct.font = Font(bold=True, color="FFFFFF"); ct.fill = PatternFill("solid", fgColor="1a3a5c")
    ct.number_format = '"$"#,##0.00'
    for i, w in enumerate([30,16,26,20,16,16,14,16], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Hoja 2: Detalle de Tickets
    ws2 = wb.create_sheet("Detalle de Tickets")
    ws2.merge_cells("A1:I1")
    ws2["A1"] = f"DETALLE — {nombre_mes.upper()} {anio}"
    ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(["Agente","Fecha","Categoría","Comprobante","Descripción","Valor","Estado","Aprobado","Motivo"], 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="2d6a9f")

    colores = {"aprobado":"D4EDDA","rechazado":"F8D7DA","debito_parcial":"FFF3CD","pendiente":"FFFFFF","pagado":"E0F2FE"}
    for ri, t in enumerate(listar_tickets(anio=anio, mes=mes), 3):
        color = colores.get(t["estado"], "FFFFFF")
        for c, v in enumerate([t["agente_nombre"],t["fecha_gasto"],t.get("categoria_nombre",""),
                                t.get("comprobante",""),t.get("descripcion",""),t["valor"],
                                t["estado"].upper(),t.get("valor_aprobado") or "",t.get("motivo_rechazo","")], 1):
            cell = ws2.cell(row=ri, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=color)
            if c in (6,8): cell.number_format = '"$"#,##0.00'
    for col, w in zip("ABCDEFGHI",[28,12,16,16,30,14,14,14,32]):
        ws2.column_dimensions[col].width = w

    # Hoja 3: Resumen Semanal
    ws3 = wb.create_sheet("Resumen Semanal")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"RESUMEN SEMANAL — {nombre_mes.upper()} {anio}"
    ws3["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    row = 2
    for s in resumen_semanal(anio, mes):
        ws3.cell(row=row, column=1, value=f"Semana {s['semana']}").font = Font(bold=True, color="FFFFFF")
        ws3.cell(row=row, column=1).fill = PatternFill("solid", fgColor="2d6a9f")
        ws3.merge_cells(f"A{row}:D{row}"); row += 1
        for c, h in enumerate(["Agente","Fecha","Descripción","Aprobado"], 1):
            ws3.cell(row=row, column=c, value=h).font = Font(bold=True)
        row += 1
        for t in s["tickets"]:
            ws3.cell(row=row,column=1,value=t["agente_nombre"])
            ws3.cell(row=row,column=2,value=t["fecha_gasto"])
            ws3.cell(row=row,column=3,value=t.get("descripcion",""))
            ws3.cell(row=row,column=4,value=t.get("valor_aprobado",0)).number_format = '"$"#,##0.00'
            row += 1
        ws3.cell(row=row,column=3,value="TOTAL SEMANA").font = Font(bold=True)
        ws3.cell(row=row,column=4,value=s["total"]).font = Font(bold=True)
        ws3.cell(row=row,column=4).number_format = '"$"#,##0.00'
        row += 2
    for col in "ABCD": ws3.column_dimensions[col].width = 28

    filename = f"viaticos_ATE_{anio}_{mes:02d}.xlsx"
    path = os.path.join(EXPORTS_DIR, filename)
    wb.save(path)
    return FileResponse(path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── EXPORTACIÓN PDF ───────────────────────────────────────────────────────────

@app.get("/api/exportar/pdf")
def exportar_pdf(anio: int, mes: int):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    resumen_data = resumen(anio, mes)
    nombre_mes   = calendar.month_name[mes]
    filename     = f"reintegros_ATE_{anio}_{mes:02d}.pdf"
    path         = os.path.join(EXPORTS_DIR, filename)

    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    AZ  = colors.HexColor("#1a3a5c")
    AM  = colors.HexColor("#2d6a9f")
    AC  = colors.HexColor("#EBF3FB")
    ts  = ParagraphStyle("t", fontSize=18, textColor=colors.white, alignment=TA_CENTER, fontName="Helvetica-Bold")
    ss  = ParagraphStyle("s", fontSize=10, textColor=colors.white, alignment=TA_CENTER, fontName="Helvetica")

    ht = Table([[Paragraph("SINDICATO ATE", ts)],
                [Paragraph(f"REINTEGRO DE VIÁTICOS — {nombre_mes.upper()} {anio}", ss)],
                [Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", ss)]],
               colWidths=[26*cm])
    ht.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),AZ),("ROWPADDING",(0,0),(-1,-1),8),
                             ("TOPPADDING",(0,0),(-1,0),14),("BOTTOMPADDING",(0,-1),(-1,-1),14)]))

    hdrs = ["Agente","CUIT","CBU","Banco/Alias","Total Presentado","Aprobado","Tope","A TRANSFERIR"]
    td = [hdrs]; total = 0
    for ag in resumen_data:
        tp = ag["total_aprobado"] + ag["total_debito"] + ag["total_pendiente"]
        ta = ag["total_aprobado"] + ag["total_debito"]
        td.append([ag["nombre"], ag.get("cuit") or "-", ag.get("cbu") or "-",
                   ag.get("banco") or ag.get("alias") or "-",
                   f"${tp:,.2f}", f"${ta:,.2f}", f"${ag['tope_efectivo']:,.2f}", f"${ag['a_transferir']:,.2f}"])
        total += ag["a_transferir"]
    td.append(["","","","TOTAL","","","", f"${total:,.2f}"])

    t = Table(td, colWidths=[5.5*cm,2.8*cm,4.2*cm,3.5*cm,2.8*cm,2.8*cm,2.3*cm,2.8*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),AM),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),9),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("FONTSIZE",(0,1),(-1,-1),8),("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white,AC]),
        ("BACKGROUND",(0,-1),(-1,-1),AZ),("TEXTCOLOR",(0,-1),(-1,-1),colors.white),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#CCCCCC")),("ROWPADDING",(0,0),(-1,-1),6),
    ]))
    doc.build([ht, Spacer(1, 0.5*cm), t])
    return FileResponse(path, filename=filename, media_type="application/pdf")


# ── BACKUP ───────────────────────────────────────────────────────────────────

@app.get("/api/backup")
def descargar_backup():
    import json
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM agentes ORDER BY id")
        agentes = _rows(cur)
        cur.execute("SELECT * FROM categorias ORDER BY id")
        categorias = _rows(cur)
        cur.execute("SELECT * FROM tickets ORDER BY id")
        tickets = _rows(cur)
        for t in tickets:
            for k in ("fecha_gasto", "creado_en", "revisado_en"):
                if t.get(k) and not isinstance(t[k], str):
                    t[k] = t[k].isoformat()
        cur.execute("SELECT * FROM periodos ORDER BY id")
        periodos = _rows(cur)
        for a in agentes:
            if a.get("creado_en") and not isinstance(a["creado_en"], str):
                a["creado_en"] = a["creado_en"].isoformat()

    backup = {
        "metadata": {
            "sistema": "Viaticos ATE",
            "version": "1.0",
            "fecha_backup": datetime.now().isoformat(),
            "motor": "postgresql" if USE_PG else "sqlite",
            "totales": {
                "agentes": len(agentes),
                "categorias": len(categorias),
                "tickets": len(tickets),
                "periodos": len(periodos),
            }
        },
        "agentes":    agentes,
        "categorias": categorias,
        "tickets":    tickets,
        "periodos":   periodos,
    }
    filename = f"backup_viaticos_ATE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    path = os.path.join(EXPORTS_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(backup, f, ensure_ascii=False, indent=2)
    return FileResponse(path, filename=filename, media_type="application/json")


@app.post("/api/restaurar")
async def restaurar_backup(request: dict):
    import json
    if "tickets" not in request or "agentes" not in request:
        raise HTTPException(400, "Archivo de backup invalido")
    with get_db() as conn:
        cur = conn.cursor()
        for tabla in ("periodos", "tickets", "agentes"):
            cur.execute(f"DELETE FROM {tabla}")
        for a in request.get("agentes", []):
            if USE_PG:
                cur.execute("""
                    INSERT INTO agentes (id,nombre,cuit,cbu,banco,alias,tope_mensual,activo,creado_en)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (id) DO UPDATE SET nombre=EXCLUDED.nombre,cuit=EXCLUDED.cuit,
                    cbu=EXCLUDED.cbu,banco=EXCLUDED.banco,alias=EXCLUDED.alias,
                    tope_mensual=EXCLUDED.tope_mensual,activo=EXCLUDED.activo
                """, (a["id"],a["nombre"],a.get("cuit"),a.get("cbu"),a.get("banco"),
                      a.get("alias"),a.get("tope_mensual",0),a.get("activo",1),a.get("creado_en")))
            else:
                cur.execute("""INSERT OR REPLACE INTO agentes
                    (id,nombre,cuit,cbu,banco,alias,tope_mensual,activo,creado_en)
                    VALUES (?,?,?,?,?,?,?,?,?)""",
                    (a["id"],a["nombre"],a.get("cuit"),a.get("cbu"),a.get("banco"),
                     a.get("alias"),a.get("tope_mensual",0),a.get("activo",1),a.get("creado_en")))
        for t in request.get("tickets", []):
            if USE_PG:
                cur.execute("""INSERT INTO tickets
                    (id,agente_id,fecha_gasto,categoria_id,comprobante,descripcion,valor,estado,
                    motivo_rechazo,valor_aprobado,creado_en,revisado_en,revisado_por)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (id) DO NOTHING""",
                    (t["id"],t["agente_id"],t["fecha_gasto"],t.get("categoria_id"),t.get("comprobante"),
                     t.get("descripcion"),t["valor"],t.get("estado","pendiente"),t.get("motivo_rechazo"),
                     t.get("valor_aprobado"),t.get("creado_en"),t.get("revisado_en"),t.get("revisado_por")))
            else:
                cur.execute("""INSERT OR REPLACE INTO tickets
                    (id,agente_id,fecha_gasto,categoria_id,comprobante,descripcion,valor,estado,
                    motivo_rechazo,valor_aprobado,creado_en,revisado_en,revisado_por)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (t["id"],t["agente_id"],t["fecha_gasto"],t.get("categoria_id"),t.get("comprobante"),
                     t.get("descripcion"),t["valor"],t.get("estado","pendiente"),t.get("motivo_rechazo"),
                     t.get("valor_aprobado"),t.get("creado_en"),t.get("revisado_en"),t.get("revisado_por")))
        for p in request.get("periodos", []):
            if USE_PG:
                cur.execute("""INSERT INTO periodos (id,agente_id,anio,mes,tope_override,cerrado)
                    VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (id) DO NOTHING""",
                    (p["id"],p["agente_id"],p["anio"],p["mes"],p.get("tope_override"),p.get("cerrado",0)))
            else:
                cur.execute("""INSERT OR REPLACE INTO periodos (id,agente_id,anio,mes,tope_override,cerrado)
                    VALUES (?,?,?,?,?,?)""",
                    (p["id"],p["agente_id"],p["anio"],p["mes"],p.get("tope_override"),p.get("cerrado",0)))
        if USE_PG:
            for tabla in ("agentes", "tickets", "periodos"):
                cur.execute(f"SELECT setval(pg_get_serial_sequence('{tabla}','id'), COALESCE(MAX(id),1)) FROM {tabla}")
    meta = request.get("metadata", {})
    return {
        "mensaje": "Backup restaurado correctamente",
        "agentes_restaurados": len(request.get("agentes", [])),
        "tickets_restaurados": len(request.get("tickets", [])),
        "fecha_backup_original": meta.get("fecha_backup", "desconocida"),
    }

# ── FRONTEND ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def frontend():
    with open(os.path.join(BASE_DIR, "templates", "index.html"), encoding="utf-8") as f:
        return f.read()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=int(os.environ.get("PORT", 8000)), reload=True)

# ── ADJUNTOS ──────────────────────────────────────────────────────────────────

from fastapi import UploadFile, File
from fastapi.responses import Response

@app.post("/api/tickets/{ticket_id}/adjunto")
async def subir_adjunto(ticket_id: int, archivo: UploadFile = File(...)):
    MAX_MB = 5
    data = await archivo.read()
    if len(data) > MAX_MB * 1024 * 1024:
        raise HTTPException(413, f"El archivo supera {MAX_MB}MB")
    allowed = {"image/jpeg","image/png","image/gif","image/webp","application/pdf"}
    mime = archivo.content_type or "application/octet-stream"
    if mime not in allowed:
        raise HTTPException(415, "Solo se permiten imágenes (JPG, PNG) o PDF")
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT id FROM tickets WHERE id={PH}", (ticket_id,))
        if not _row(cur):
            raise HTTPException(404, "Ticket no encontrado")
        if USE_PG:
            import psycopg2
            cur.execute(
                "INSERT INTO adjuntos (ticket_id,nombre_archivo,mime_type,datos) VALUES (%s,%s,%s,%s) RETURNING id",
                (ticket_id, archivo.filename, mime, psycopg2.Binary(data))
            )
            new_id = cur.fetchone()[0]
        else:
            cur.execute(
                "INSERT INTO adjuntos (ticket_id,nombre_archivo,mime_type,datos) VALUES (?,?,?,?)",
                (ticket_id, archivo.filename, mime, data)
            )
            new_id = cur.lastrowid
    return {"id": new_id, "nombre": archivo.filename, "mime_type": mime, "size": len(data)}

@app.get("/api/tickets/{ticket_id}/adjuntos")
def listar_adjuntos(ticket_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT id,nombre_archivo,mime_type,subido_en FROM adjuntos WHERE ticket_id={PH} ORDER BY subido_en", (ticket_id,))
        rows = _rows(cur)
        for r in rows:
            if r.get("subido_en") and not isinstance(r["subido_en"], str):
                r["subido_en"] = r["subido_en"].isoformat()
        return rows

@app.get("/api/adjuntos/{adjunto_id}")
def descargar_adjunto(adjunto_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM adjuntos WHERE id={PH}", (adjunto_id,))
        row = _row(cur)
        if not row:
            raise HTTPException(404, "Adjunto no encontrado")
    data = bytes(row["datos"]) if not isinstance(row["datos"], bytes) else row["datos"]
    return Response(content=data, media_type=row["mime_type"],
                    headers={"Content-Disposition": f'inline; filename="{row["nombre_archivo"]}"'})

@app.delete("/api/adjuntos/{adjunto_id}")
def eliminar_adjunto(adjunto_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM adjuntos WHERE id={PH}", (adjunto_id,))
    return {"mensaje": "Adjunto eliminado"}

# ── HISTORIAL ─────────────────────────────────────────────────────────────────

@app.get("/api/tickets/{ticket_id}/historial")
def obtener_historial(ticket_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM historial WHERE ticket_id={PH} ORDER BY fecha DESC", (ticket_id,))
        rows = _rows(cur)
        for r in rows:
            if r.get("fecha") and not isinstance(r["fecha"], str):
                r["fecha"] = r["fecha"].isoformat()
        return rows

def _registrar_historial(conn, ticket_id, campo, anterior, nuevo, usuario=None):
    cur = conn.cursor()
    if USE_PG:
        cur.execute(
            "INSERT INTO historial (ticket_id,campo,valor_anterior,valor_nuevo,usuario) VALUES (%s,%s,%s,%s,%s)",
            (ticket_id, campo, str(anterior) if anterior is not None else None,
             str(nuevo) if nuevo is not None else None, usuario)
        )
    else:
        cur.execute(
            "INSERT INTO historial (ticket_id,campo,valor_anterior,valor_nuevo,usuario) VALUES (?,?,?,?,?)",
            (ticket_id, campo, str(anterior) if anterior is not None else None,
             str(nuevo) if nuevo is not None else None, usuario)
        )

# ── CIERRE DE PERÍODO ─────────────────────────────────────────────────────────

@app.get("/api/periodos_cierre")
def listar_cierres():
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM periodos_cierre ORDER BY anio DESC, mes DESC")
        rows = _rows(cur)
        for r in rows:
            if r.get("cerrado_en") and not isinstance(r["cerrado_en"], str):
                r["cerrado_en"] = r["cerrado_en"].isoformat()
        return rows

@app.post("/api/periodos_cierre")
def cerrar_periodo(anio: int, mes: int, usuario: Optional[str] = None):
    with get_db() as conn:
        cur = conn.cursor()
        # Verificar que no haya tickets pendientes
        cur.execute(f"""SELECT COUNT(*) as n FROM tickets
                    WHERE estado='pendiente' AND {_year_filter('fecha_gasto')} AND {_month_filter('fecha_gasto')}""",
                    (_yp(anio), _mp(mes)))
        row = _row(cur)
        pendientes = row["n"] if row else 0
        if pendientes > 0:
            raise HTTPException(409, f"Hay {pendientes} ticket(s) pendientes de revisión. Revisalos antes de cerrar el período.")
        if USE_PG:
            cur.execute("""
                INSERT INTO periodos_cierre (anio,mes,cerrado,cerrado_por,cerrado_en)
                VALUES (%s,%s,1,%s,CURRENT_TIMESTAMP)
                ON CONFLICT (anio,mes) DO UPDATE SET cerrado=1, cerrado_por=EXCLUDED.cerrado_por, cerrado_en=CURRENT_TIMESTAMP
            """, (anio, mes, usuario))
        else:
            cur.execute("""
                INSERT INTO periodos_cierre (anio,mes,cerrado,cerrado_por,cerrado_en)
                VALUES (?,?,1,?,CURRENT_TIMESTAMP)
                ON CONFLICT(anio,mes) DO UPDATE SET cerrado=1, cerrado_por=excluded.cerrado_por, cerrado_en=CURRENT_TIMESTAMP
            """, (anio, mes, usuario))
    return {"mensaje": f"Período {mes}/{anio} cerrado correctamente"}

@app.post("/api/periodos_cierre/reabrir")
def reabrir_periodo(anio: int, mes: int, usuario: Optional[str] = None):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"UPDATE periodos_cierre SET cerrado=0, cerrado_por={PH}, cerrado_en=CURRENT_TIMESTAMP WHERE anio={PH} AND mes={PH}",
                    (usuario, anio, mes))
    return {"mensaje": f"Período {mes}/{anio} reabierto"}

@app.get("/api/periodos_cierre/estado")
def estado_periodo(anio: int, mes: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM periodos_cierre WHERE anio={PH} AND mes={PH}", (anio, mes))
        row = _row(cur)
        return {"cerrado": bool(row and row.get("cerrado")), "detalle": row}

# ── REPORTES ──────────────────────────────────────────────────────────────────

@app.get("/api/reportes/por_categoria")
def reporte_por_categoria(anio: Optional[int] = None, mes: Optional[int] = None):
    with get_db() as conn:
        cur = conn.cursor()
        q = """SELECT c.nombre as categoria,
                      COUNT(*) as cantidad,
                      SUM(t.valor) as total_presentado,
                      SUM(CASE WHEN t.estado IN ('aprobado','debito_parcial','pagado') THEN t.valor_aprobado ELSE 0 END) as total_aprobado,
                      COUNT(CASE WHEN t.estado='aprobado' THEN 1 END) as cant_aprobados,
                      COUNT(CASE WHEN t.estado='rechazado' THEN 1 END) as cant_rechazados
               FROM tickets t
               LEFT JOIN categorias c ON t.categoria_id = c.id
               WHERE 1=1"""
        params = []
        if anio:
            q += f" AND {_year_filter('t.fecha_gasto')}"; params.append(_yp(anio))
        if mes:
            q += f" AND {_month_filter('t.fecha_gasto')}"; params.append(_mp(mes))
        q += " GROUP BY c.nombre ORDER BY total_aprobado DESC"
        cur.execute(q, params)
        return _rows(cur)

@app.get("/api/reportes/mensual_anual")
def reporte_mensual_anual(anio: int):
    with get_db() as conn:
        cur = conn.cursor()
        if USE_PG:
            cur.execute("""
                SELECT EXTRACT(MONTH FROM fecha_gasto)::int as mes,
                       COUNT(*) as cantidad,
                       SUM(valor) as total_presentado,
                       SUM(CASE WHEN estado IN ('aprobado','debito_parcial','pagado') THEN valor_aprobado ELSE 0 END) as total_aprobado,
                       COUNT(CASE WHEN estado='pendiente' THEN 1 END) as pendientes
                FROM tickets WHERE EXTRACT(YEAR FROM fecha_gasto) = %s
                GROUP BY mes ORDER BY mes
            """, (anio,))
        else:
            cur.execute("""
                SELECT CAST(strftime('%m', fecha_gasto) AS INTEGER) as mes,
                       COUNT(*) as cantidad,
                       SUM(valor) as total_presentado,
                       SUM(CASE WHEN estado IN ('aprobado','debito_parcial','pagado') THEN valor_aprobado ELSE 0 END) as total_aprobado,
                       COUNT(CASE WHEN estado='pendiente' THEN 1 END) as pendientes
                FROM tickets WHERE strftime('%Y', fecha_gasto) = ?
                GROUP BY mes ORDER BY mes
            """, (str(anio),))
        return _rows(cur)

@app.get("/api/reportes/por_agente_anual")
def reporte_por_agente_anual(anio: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT a.nombre,
                   COUNT(t.id) as total_tickets,
                   SUM(t.valor) as total_presentado,
                   SUM(CASE WHEN t.estado IN ('aprobado','debito_parcial','pagado') THEN t.valor_aprobado ELSE 0 END) as total_aprobado,
                   COUNT(CASE WHEN t.estado='rechazado' THEN 1 END) as rechazados,
                   COUNT(CASE WHEN t.estado='pendiente' THEN 1 END) as pendientes
            FROM agentes a
            LEFT JOIN tickets t ON a.id = t.agente_id AND {_year_filter('t.fecha_gasto')}
            WHERE a.activo = 1
            GROUP BY a.id, a.nombre ORDER BY total_aprobado DESC NULLS LAST
        """, (_yp(anio),))
        return _rows(cur)

# ── ADJUNTOS ──────────────────────────────────────────────────────────────────

from fastapi import UploadFile, File
from fastapi.responses import Response

@app.post("/api/tickets/{ticket_id}/adjunto")
async def subir_adjunto(ticket_id: int, archivo: UploadFile = File(...)):
    MAX_MB = 5
    data = await archivo.read()
    if len(data) > MAX_MB * 1024 * 1024:
        raise HTTPException(413, f"El archivo supera {MAX_MB}MB")
    allowed = {"image/jpeg","image/png","image/gif","image/webp","application/pdf"}
    mime = archivo.content_type or "application/octet-stream"
    if mime not in allowed:
        raise HTTPException(415, "Solo se permiten imagenes (JPG, PNG) o PDF")
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT id FROM tickets WHERE id={PH}", (ticket_id,))
        if not _row(cur):
            raise HTTPException(404, "Ticket no encontrado")
        if USE_PG:
            import psycopg2
            cur.execute(
                "INSERT INTO adjuntos (ticket_id,nombre_archivo,mime_type,datos) VALUES (%s,%s,%s,%s) RETURNING id",
                (ticket_id, archivo.filename, mime, psycopg2.Binary(data))
            )
            new_id = cur.fetchone()[0]
        else:
            cur.execute(
                "INSERT INTO adjuntos (ticket_id,nombre_archivo,mime_type,datos) VALUES (?,?,?,?)",
                (ticket_id, archivo.filename, mime, data)
            )
            new_id = cur.lastrowid
    return {"id": new_id, "nombre": archivo.filename, "mime_type": mime, "size": len(data)}

@app.get("/api/tickets/{ticket_id}/adjuntos")
def listar_adjuntos(ticket_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT id,nombre_archivo,mime_type,subido_en FROM adjuntos WHERE ticket_id={PH} ORDER BY subido_en", (ticket_id,))
        rows = _rows(cur)
        for r in rows:
            if r.get("subido_en") and not isinstance(r["subido_en"], str):
                r["subido_en"] = r["subido_en"].isoformat()
        return rows

@app.get("/api/adjuntos/{adjunto_id}")
def descargar_adjunto(adjunto_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM adjuntos WHERE id={PH}", (adjunto_id,))
        row = _row(cur)
        if not row:
            raise HTTPException(404, "Adjunto no encontrado")
    data = bytes(row["datos"]) if not isinstance(row["datos"], bytes) else row["datos"]
    return Response(content=data, media_type=row["mime_type"],
                    headers={"Content-Disposition": f'inline; filename="{row["nombre_archivo"]}"'})

@app.delete("/api/adjuntos/{adjunto_id}")
def eliminar_adjunto(adjunto_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM adjuntos WHERE id={PH}", (adjunto_id,))
    return {"mensaje": "Adjunto eliminado"}

# ── HISTORIAL ─────────────────────────────────────────────────────────────────

@app.get("/api/tickets/{ticket_id}/historial")
def obtener_historial(ticket_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM historial WHERE ticket_id={PH} ORDER BY fecha DESC", (ticket_id,))
        rows = _rows(cur)
        for r in rows:
            if r.get("fecha") and not isinstance(r["fecha"], str):
                r["fecha"] = r["fecha"].isoformat()
        return rows

def _registrar_historial(conn, ticket_id, campo, anterior, nuevo, usuario=None):
    cur = conn.cursor()
    if USE_PG:
        cur.execute(
            "INSERT INTO historial (ticket_id,campo,valor_anterior,valor_nuevo,usuario) VALUES (%s,%s,%s,%s,%s)",
            (ticket_id, campo, str(anterior) if anterior is not None else None,
             str(nuevo) if nuevo is not None else None, usuario)
        )
    else:
        cur.execute(
            "INSERT INTO historial (ticket_id,campo,valor_anterior,valor_nuevo,usuario) VALUES (?,?,?,?,?)",
            (ticket_id, campo, str(anterior) if anterior is not None else None,
             str(nuevo) if nuevo is not None else None, usuario)
        )

# ── CIERRE DE PERIODO ─────────────────────────────────────────────────────────

@app.get("/api/periodos_cierre/estado")
def estado_periodo(anio: int, mes: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM periodos_cierre WHERE anio={PH} AND mes={PH}", (anio, mes))
        row = _row(cur)
        return {"cerrado": bool(row and row.get("cerrado")), "detalle": row}

@app.get("/api/periodos_cierre")
def listar_cierres():
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM periodos_cierre ORDER BY anio DESC, mes DESC")
        rows = _rows(cur)
        for r in rows:
            if r.get("cerrado_en") and not isinstance(r["cerrado_en"], str):
                r["cerrado_en"] = r["cerrado_en"].isoformat()
        return rows

@app.post("/api/periodos_cierre")
def cerrar_periodo(anio: int, mes: int, usuario: Optional[str] = None):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"""SELECT COUNT(*) as n FROM tickets
                    WHERE estado='pendiente' AND {_year_filter('fecha_gasto')} AND {_month_filter('fecha_gasto')}""",
                    (_yp(anio), _mp(mes)))
        row = _row(cur)
        pendientes = row["n"] if row else 0
        if pendientes > 0:
            raise HTTPException(409, f"Hay {pendientes} ticket(s) pendientes de revision. Revisalos antes de cerrar el periodo.")
        if USE_PG:
            cur.execute("""
                INSERT INTO periodos_cierre (anio,mes,cerrado,cerrado_por,cerrado_en)
                VALUES (%s,%s,1,%s,CURRENT_TIMESTAMP)
                ON CONFLICT (anio,mes) DO UPDATE SET cerrado=1,cerrado_por=EXCLUDED.cerrado_por,cerrado_en=CURRENT_TIMESTAMP
            """, (anio, mes, usuario))
        else:
            cur.execute("""
                INSERT INTO periodos_cierre (anio,mes,cerrado,cerrado_por,cerrado_en)
                VALUES (?,?,1,?,CURRENT_TIMESTAMP)
                ON CONFLICT(anio,mes) DO UPDATE SET cerrado=1,cerrado_por=excluded.cerrado_por,cerrado_en=CURRENT_TIMESTAMP
            """, (anio, mes, usuario))
    return {"mensaje": f"Periodo {mes}/{anio} cerrado correctamente"}

@app.post("/api/periodos_cierre/reabrir")
def reabrir_periodo(anio: int, mes: int, usuario: Optional[str] = None):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"UPDATE periodos_cierre SET cerrado=0,cerrado_por={PH},cerrado_en=CURRENT_TIMESTAMP WHERE anio={PH} AND mes={PH}",
                    (usuario, anio, mes))
    return {"mensaje": f"Periodo {mes}/{anio} reabierto"}

# ── REPORTES ──────────────────────────────────────────────────────────────────

@app.get("/api/reportes/por_categoria")
def reporte_por_categoria(anio: Optional[int] = None, mes: Optional[int] = None):
    with get_db() as conn:
        cur = conn.cursor()
        q = """SELECT COALESCE(c.nombre,'Sin categoria') as categoria,
                      COUNT(*) as cantidad,
                      SUM(t.valor) as total_presentado,
                      SUM(CASE WHEN t.estado IN ('aprobado','debito_parcial','pagado') THEN t.valor_aprobado ELSE 0 END) as total_aprobado,
                      COUNT(CASE WHEN t.estado='aprobado' THEN 1 END) as cant_aprobados,
                      COUNT(CASE WHEN t.estado='rechazado' THEN 1 END) as cant_rechazados
               FROM tickets t LEFT JOIN categorias c ON t.categoria_id = c.id WHERE 1=1"""
        params = []
        if anio:
            q += f" AND {_year_filter('t.fecha_gasto')}"; params.append(_yp(anio))
        if mes:
            q += f" AND {_month_filter('t.fecha_gasto')}"; params.append(_mp(mes))
        q += " GROUP BY c.nombre ORDER BY total_aprobado DESC"
        cur.execute(q, params)
        return _rows(cur)

@app.get("/api/reportes/mensual_anual")
def reporte_mensual_anual(anio: int):
    with get_db() as conn:
        cur = conn.cursor()
        if USE_PG:
            cur.execute("""
                SELECT EXTRACT(MONTH FROM fecha_gasto)::int as mes,
                       COUNT(*) as cantidad, SUM(valor) as total_presentado,
                       SUM(CASE WHEN estado IN ('aprobado','debito_parcial','pagado') THEN valor_aprobado ELSE 0 END) as total_aprobado,
                       COUNT(CASE WHEN estado='pendiente' THEN 1 END) as pendientes
                FROM tickets WHERE EXTRACT(YEAR FROM fecha_gasto)=%s
                GROUP BY mes ORDER BY mes
            """, (anio,))
        else:
            cur.execute("""
                SELECT CAST(strftime('%m',fecha_gasto) AS INTEGER) as mes,
                       COUNT(*) as cantidad, SUM(valor) as total_presentado,
                       SUM(CASE WHEN estado IN ('aprobado','debito_parcial','pagado') THEN valor_aprobado ELSE 0 END) as total_aprobado,
                       COUNT(CASE WHEN estado='pendiente' THEN 1 END) as pendientes
                FROM tickets WHERE strftime('%Y',fecha_gasto)=?
                GROUP BY mes ORDER BY mes
            """, (str(anio),))
        return _rows(cur)

@app.get("/api/reportes/por_agente_anual")
def reporte_por_agente_anual(anio: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT a.nombre,
                   COUNT(t.id) as total_tickets,
                   SUM(COALESCE(t.valor,0)) as total_presentado,
                   SUM(CASE WHEN t.estado IN ('aprobado','debito_parcial','pagado') THEN COALESCE(t.valor_aprobado,0) ELSE 0 END) as total_aprobado,
                   COUNT(CASE WHEN t.estado='rechazado' THEN 1 END) as rechazados,
                   COUNT(CASE WHEN t.estado='pendiente' THEN 1 END) as pendientes
            FROM agentes a
            LEFT JOIN tickets t ON a.id=t.agente_id AND {_year_filter('t.fecha_gasto')}
            WHERE a.activo=1
            GROUP BY a.id, a.nombre ORDER BY total_aprobado DESC
        """, (_yp(anio),))
        return _rows(cur)

# ── ACTAS ─────────────────────────────────────────────────────────────────────

TIPOS_ACTA = ["Entrega de Viáticos", "Reunión / Asamblea", "Llamado / Citación", "Resolución / Acuerdo", "Otro"]

class ActaCreate(BaseModel):
    fecha: str
    tipo: str
    titulo: str
    cuerpo: Optional[str] = None
    participantes: Optional[str] = None
    redactado_por: Optional[str] = None

class ActaUpdate(BaseModel):
    fecha: Optional[str] = None
    tipo: Optional[str] = None
    titulo: Optional[str] = None
    cuerpo: Optional[str] = None
    participantes: Optional[str] = None
    redactado_por: Optional[str] = None

def _acta_to_dict(r):
    d = dict(r)
    for k in ("fecha", "creado_en", "modificado_en"):
        if d.get(k) and not isinstance(d[k], str):
            d[k] = d[k].isoformat()
    return d

def _gen_numero_acta(conn):
    """Genera número correlativo tipo 0001/2026"""
    anio = datetime.now().year
    cur = conn.cursor()
    if USE_PG:
        cur.execute("SELECT COUNT(*) as n FROM actas WHERE EXTRACT(YEAR FROM fecha::date) = %s", (anio,))
    else:
        cur.execute("SELECT COUNT(*) as n FROM actas WHERE strftime('%Y', fecha) = ?", (str(anio),))
    row = _row(cur)
    n = (row["n"] if row else 0) + 1
    return f"{n:04d}/{anio}"

@app.get("/api/actas")
def listar_actas(buscar: Optional[str] = None, tipo: Optional[str] = None,
                 anio: Optional[int] = None, mes: Optional[int] = None):
    with get_db() as conn:
        cur = conn.cursor()
        q = "SELECT * FROM actas WHERE 1=1"
        params = []
        if tipo:
            q += f" AND tipo = {PH}"; params.append(tipo)
        if anio:
            if USE_PG:
                q += f" AND EXTRACT(YEAR FROM fecha::date) = {PH}"
            else:
                q += f" AND strftime('%Y', fecha) = {PH}"
            params.append(_yp(anio))
        if mes:
            if USE_PG:
                q += f" AND EXTRACT(MONTH FROM fecha::date) = {PH}"
            else:
                q += f" AND strftime('%m', fecha) = {PH}"
            params.append(_mp(mes))
        if buscar:
            if USE_PG:
                q += f" AND (titulo ILIKE {PH} OR cuerpo ILIKE {PH} OR participantes ILIKE {PH} OR numero_acta ILIKE {PH})"
                p = f"%{buscar}%"
                params += [p, p, p, p]
            else:
                q += f" AND (titulo LIKE {PH} OR cuerpo LIKE {PH} OR participantes LIKE {PH} OR numero_acta LIKE {PH})"
                p = f"%{buscar}%"
                params += [p, p, p, p]
        q += " ORDER BY fecha DESC, id DESC"
        cur.execute(q, params)
        return [_acta_to_dict(r) for r in _rows(cur)]

@app.get("/api/actas/{acta_id}")
def obtener_acta(acta_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM actas WHERE id = {PH}", (acta_id,))
        row = _row(cur)
        if not row:
            raise HTTPException(404, "Acta no encontrada")
        return _acta_to_dict(row)

@app.post("/api/actas")
def crear_acta(data: ActaCreate):
    with get_db() as conn:
        numero = _gen_numero_acta(conn)
        if USE_PG:
            new_id = _insert(conn,
                "INSERT INTO actas (numero_acta,fecha,tipo,titulo,cuerpo,participantes,redactado_por) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (numero, data.fecha, data.tipo, data.titulo, data.cuerpo, data.participantes, data.redactado_por))
        else:
            new_id = _insert(conn,
                "INSERT INTO actas (numero_acta,fecha,tipo,titulo,cuerpo,participantes,redactado_por) VALUES (?,?,?,?,?,?,?)",
                (numero, data.fecha, data.tipo, data.titulo, data.cuerpo, data.participantes, data.redactado_por))
    return {"id": new_id, "numero_acta": numero, "mensaje": "Acta creada"}

@app.put("/api/actas/{acta_id}")
def actualizar_acta(acta_id: int, data: ActaUpdate):
    fields = {k: v for k, v in data.dict().items() if v is not None}
    if not fields:
        raise HTTPException(400, "Nada para actualizar")
    fields["modificado_en"] = datetime.now().isoformat()
    with get_db() as conn:
        cur = conn.cursor()
        if USE_PG:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            cur.execute(f"UPDATE actas SET {set_clause} WHERE id=%s", list(fields.values()) + [acta_id])
        else:
            set_clause = ", ".join(f"{k}=?" for k in fields)
            cur.execute(f"UPDATE actas SET {set_clause} WHERE id=?", list(fields.values()) + [acta_id])
    return {"mensaje": "Acta actualizada"}

@app.delete("/api/actas/{acta_id}")
def eliminar_acta(acta_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM actas WHERE id = {PH}", (acta_id,))
    return {"mensaje": "Acta eliminada"}

@app.get("/api/actas/exportar/excel")
def exportar_actas_excel(anio: Optional[int] = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro de Actas"

    titulo = f"SINDICATO ATE — LIBRO DE ACTAS{f' — {anio}' if anio else ''}"
    ws.merge_cells("A1:H1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdrs = ["N° Acta", "Fecha", "Tipo", "Título", "Participantes", "Cuerpo / Contenido", "Redactado por", "Registrado"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2d6a9f")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    tipo_colores = {
        "Entrega de Viáticos":  "D4EDDA",
        "Reunión / Asamblea":   "D1ECF1",
        "Llamado / Citación":   "FFF3CD",
        "Resolución / Acuerdo": "EDE7F6",
        "Otro":                 "F8F9FA",
    }

    actas = listar_actas(anio=anio)
    for ri, a in enumerate(actas, 3):
        color = tipo_colores.get(a.get("tipo", ""), "FFFFFF")
        fecha_fmt = ""
        if a.get("fecha"):
            try:
                from datetime import date
                fd = a["fecha"][:10]
                parts = fd.split("-")
                fecha_fmt = f"{parts[2]}/{parts[1]}/{parts[0]}"
            except:
                fecha_fmt = a.get("fecha", "")

        vals = [
            a.get("numero_acta", ""),
            fecha_fmt,
            a.get("tipo", ""),
            a.get("titulo", ""),
            a.get("participantes", "") or "",
            a.get("cuerpo", "") or "",
            a.get("redactado_por", "") or "",
            (a.get("creado_en", "") or "")[:10],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(wrap_text=(c in (5, 6)), vertical="top")

    col_widths = [12, 12, 22, 36, 30, 60, 18, 12]
    for col, w in zip("ABCDEFGH", col_widths):
        ws.column_dimensions[col].width = w

    # Auto-height for body rows
    for row in ws.iter_rows(min_row=3, max_row=len(actas)+2):
        ws.row_dimensions[row[0].row].height = 40

    filename = f"libro_actas_ATE{'_'+str(anio) if anio else ''}.xlsx"
    path = os.path.join(EXPORTS_DIR, filename)
    wb.save(path)
    return FileResponse(path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── ACTAS ─────────────────────────────────────────────────────────────────────

TIPOS_ACTA = ["Entrega de Viaticos", "Reunion / Asamblea", "Llamado / Citacion", "Resolucion / Acuerdo", "Otro"]

class ActaCreate(BaseModel):
    fecha: str
    tipo: str
    titulo: str
    cuerpo: Optional[str] = None
    participantes: Optional[str] = None
    redactado_por: Optional[str] = None

class ActaUpdate(BaseModel):
    fecha: Optional[str] = None
    tipo: Optional[str] = None
    titulo: Optional[str] = None
    cuerpo: Optional[str] = None
    participantes: Optional[str] = None
    redactado_por: Optional[str] = None

def _acta_to_dict(r):
    d = dict(r)
    for k in ("fecha", "creado_en", "modificado_en"):
        if d.get(k) and not isinstance(d[k], str):
            d[k] = d[k].isoformat()
    return d

def _gen_numero_acta(conn):
    anio = datetime.now().year
    cur = conn.cursor()
    if USE_PG:
        cur.execute("SELECT COUNT(*) as n FROM actas WHERE EXTRACT(YEAR FROM fecha::date) = %s", (anio,))
    else:
        cur.execute("SELECT COUNT(*) as n FROM actas WHERE strftime('%Y', fecha) = ?", (str(anio),))
    row = _row(cur)
    n = (row["n"] if row else 0) + 1
    return f"{n:04d}/{anio}"

@app.get("/api/actas")
def listar_actas(buscar: Optional[str] = None, tipo: Optional[str] = None,
                 anio: Optional[int] = None, mes: Optional[int] = None):
    with get_db() as conn:
        cur = conn.cursor()
        q = "SELECT * FROM actas WHERE 1=1"
        params = []
        if tipo:
            q += f" AND tipo = {PH}"; params.append(tipo)
        if anio:
            if USE_PG:
                q += " AND EXTRACT(YEAR FROM fecha::date) = %s"
            else:
                q += " AND strftime('%Y', fecha) = ?"
            params.append(_yp(anio))
        if mes:
            if USE_PG:
                q += " AND EXTRACT(MONTH FROM fecha::date) = %s"
            else:
                q += " AND strftime('%m', fecha) = ?"
            params.append(_mp(mes))
        if buscar:
            p = f"%{buscar}%"
            if USE_PG:
                q += " AND (titulo ILIKE %s OR cuerpo ILIKE %s OR participantes ILIKE %s OR numero_acta ILIKE %s)"
                params += [p, p, p, p]
            else:
                q += " AND (titulo LIKE ? OR cuerpo LIKE ? OR participantes LIKE ? OR numero_acta LIKE ?)"
                params += [p, p, p, p]
        q += " ORDER BY fecha DESC, id DESC"
        cur.execute(q, params)
        return [_acta_to_dict(r) for r in _rows(cur)]

@app.get("/api/actas/{acta_id}")
def obtener_acta(acta_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM actas WHERE id = {PH}", (acta_id,))
        row = _row(cur)
        if not row:
            raise HTTPException(404, "Acta no encontrada")
        return _acta_to_dict(row)

@app.post("/api/actas")
def crear_acta(data: ActaCreate):
    with get_db() as conn:
        numero = _gen_numero_acta(conn)
        if USE_PG:
            new_id = _insert(conn,
                "INSERT INTO actas (numero_acta,fecha,tipo,titulo,cuerpo,participantes,redactado_por) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (numero, data.fecha, data.tipo, data.titulo, data.cuerpo, data.participantes, data.redactado_por))
        else:
            new_id = _insert(conn,
                "INSERT INTO actas (numero_acta,fecha,tipo,titulo,cuerpo,participantes,redactado_por) VALUES (?,?,?,?,?,?,?)",
                (numero, data.fecha, data.tipo, data.titulo, data.cuerpo, data.participantes, data.redactado_por))
    return {"id": new_id, "numero_acta": numero, "mensaje": "Acta creada"}

@app.put("/api/actas/{acta_id}")
def actualizar_acta(acta_id: int, data: ActaUpdate):
    fields = {k: v for k, v in data.dict().items() if v is not None}
    if not fields:
        raise HTTPException(400, "Nada para actualizar")
    fields["modificado_en"] = datetime.now().isoformat()
    with get_db() as conn:
        cur = conn.cursor()
        if USE_PG:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            cur.execute(f"UPDATE actas SET {set_clause} WHERE id=%s", list(fields.values()) + [acta_id])
        else:
            set_clause = ", ".join(f"{k}=?" for k in fields)
            cur.execute(f"UPDATE actas SET {set_clause} WHERE id=?", list(fields.values()) + [acta_id])
    return {"mensaje": "Acta actualizada"}

@app.delete("/api/actas/{acta_id}")
def eliminar_acta(acta_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM actas WHERE id = {PH}", (acta_id,))
    return {"mensaje": "Acta eliminada"}

@app.get("/api/actas/exportar/excel")
def exportar_actas_excel(anio: Optional[int] = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment as XlAlign
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro de Actas"
    titulo_hdr = f"SINDICATO ATE - LIBRO DE ACTAS{' - '+str(anio) if anio else ''}"
    ws.merge_cells("A1:H1")
    ws["A1"] = titulo_hdr
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws["A1"].alignment = XlAlign(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    hdrs = ["N Acta", "Fecha", "Tipo", "Titulo", "Participantes", "Contenido", "Redactado por", "Registrado"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2d6a9f")
        cell.alignment = XlAlign(horizontal="center")
    ws.row_dimensions[2].height = 20
    tipo_colores = {
        "Entrega de Viaticos":   "D4EDDA",
        "Reunion / Asamblea":    "D1ECF1",
        "Llamado / Citacion":    "FFF3CD",
        "Resolucion / Acuerdo":  "EDE7F6",
        "Otro":                  "F8F9FA",
    }
    actas_data = listar_actas(anio=anio)
    for ri, a in enumerate(actas_data, 3):
        color = tipo_colores.get(a.get("tipo", ""), "FFFFFF")
        fecha_fmt = ""
        if a.get("fecha"):
            try:
                fd = a["fecha"][:10].split("-")
                fecha_fmt = f"{fd[2]}/{fd[1]}/{fd[0]}"
            except:
                fecha_fmt = a.get("fecha", "")
        vals = [a.get("numero_acta",""), fecha_fmt, a.get("tipo",""), a.get("titulo",""),
                a.get("participantes","") or "", a.get("cuerpo","") or "",
                a.get("redactado_por","") or "", (a.get("creado_en","") or "")[:10]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = XlAlign(wrap_text=(c in (5,6)), vertical="top")
        ws.row_dimensions[ri].height = 40
    for col, w in zip("ABCDEFGH", [12,12,22,36,30,60,18,12]):
        ws.column_dimensions[col].width = w
    filename = f"libro_actas_ATE{'_'+str(anio) if anio else ''}.xlsx"
    path = os.path.join(EXPORTS_DIR, filename)
    wb.save(path)
    return FileResponse(path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/resumen/agrupado")
def resumen_agrupado(anio: Optional[int] = None):
    """Resumen por agente para año completo o todo el historial. Sin topes (no aplica)."""
    with get_db() as conn:
        cur = conn.cursor()
        q = f"""
            SELECT a.id, a.nombre, a.tope_mensual, a.cbu, a.banco, a.alias, a.cuit,
                COALESCE(SUM(CASE WHEN t.estado='pendiente'      THEN t.valor          END), 0) as total_pendiente,
                COALESCE(SUM(CASE WHEN t.estado='aprobado'       THEN t.valor_aprobado END), 0) as total_aprobado,
                COALESCE(SUM(CASE WHEN t.estado='debito_parcial' THEN t.valor_aprobado END), 0) as total_debito,
                COALESCE(SUM(CASE WHEN t.estado='pagado'         THEN t.valor_aprobado END), 0) as total_pagado,
                COALESCE(SUM(CASE WHEN t.estado='rechazado'      THEN t.valor          END), 0) as total_rechazado,
                COUNT(CASE WHEN t.estado='pendiente'      THEN 1 END) as cant_pendientes,
                COUNT(CASE WHEN t.estado='aprobado'       THEN 1 END) as cant_aprobados,
                COUNT(CASE WHEN t.estado='debito_parcial' THEN 1 END) as cant_debito,
                COUNT(CASE WHEN t.estado='pagado'         THEN 1 END) as cant_pagados,
                COUNT(CASE WHEN t.estado='rechazado'      THEN 1 END) as cant_rechazados
            FROM agentes a
            LEFT JOIN tickets t ON a.id = t.agente_id"""
        params = []
        if anio:
            q += f" AND {_year_filter('t.fecha_gasto')}"
            params.append(_yp(anio))
        q += " WHERE a.activo = 1 GROUP BY a.id, a.nombre, a.tope_mensual, a.cbu, a.banco, a.alias, a.cuit ORDER BY a.nombre"
        cur.execute(q, params)
        rows = _rows(cur)
        result = []
        for d in rows:
            subtotal = d["total_aprobado"] + d["total_debito"] + d["total_pagado"]
            d["tope_efectivo"] = 0  # no aplica tope en vista anual/histórica
            d["a_transferir"]  = subtotal
            d["excedente"]     = 0
            result.append(d)
        return result

# ── PERIODOS DE PAGO ──────────────────────────────────────────────────────────

class PeriodoPagoCreate(BaseModel):
    nombre: str
    descripcion: Optional[str] = None

class PeriodoPagoUpdate(BaseModel):
    nombre: Optional[str] = None
    descripcion: Optional[str] = None

def _get_periodo_abierto(conn):
    cur = conn.cursor()
    cur.execute("SELECT * FROM periodos_pago WHERE estado='abierto' ORDER BY numero DESC LIMIT 1")
    return _row(cur)

def _periodo_to_dict(r):
    d = dict(r)
    for k in ("creado_en", "cerrado_en"):
        if d.get(k) and not isinstance(d[k], str):
            d[k] = d[k].isoformat()
    return d

@app.get("/api/periodos_pago")
def listar_periodos_pago():
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT pp.*,
                COUNT(t.id) as total_tickets,
                COALESCE(SUM(CASE WHEN t.estado='pendiente' THEN 1 END),0) as cant_pendientes,
                COALESCE(SUM(CASE WHEN t.estado='aprobado' THEN 1 END),0) as cant_aprobados,
                COALESCE(SUM(CASE WHEN t.estado='debito_parcial' THEN 1 END),0) as cant_debito,
                COALESCE(SUM(CASE WHEN t.estado='pagado' THEN 1 END),0) as cant_pagados,
                COALESCE(SUM(CASE WHEN t.estado='rechazado' THEN 1 END),0) as cant_rechazados,
                COALESCE(SUM(CASE WHEN t.estado IN ('aprobado','debito_parcial','pagado') THEN t.valor_aprobado END),0) as total_aprobado
            FROM periodos_pago pp
            LEFT JOIN tickets t ON t.periodo_pago_id = pp.id
            GROUP BY pp.id ORDER BY pp.numero DESC
        """)
        return [_periodo_to_dict(r) for r in _rows(cur)]

@app.get("/api/periodos_pago/activo")
def get_periodo_activo():
    with get_db() as conn:
        pp = _get_periodo_abierto(conn)
        return _periodo_to_dict(pp) if pp else None

@app.post("/api/periodos_pago")
def crear_periodo_pago(data: PeriodoPagoCreate):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COALESCE(MAX(numero),0)+1 as n FROM periodos_pago")
        numero = _row(cur)["n"]
        if USE_PG:
            cur.execute(
                "INSERT INTO periodos_pago (numero,nombre,descripcion,estado) VALUES (%s,%s,%s,%s) RETURNING id",
                (numero, data.nombre, data.descripcion, 'abierto'))
            new_id = cur.fetchone()[0]
        else:
            cur.execute(
                "INSERT INTO periodos_pago (numero,nombre,descripcion,estado) VALUES (?,?,?,?)",
                (numero, data.nombre, data.descripcion, 'abierto'))
            new_id = cur.lastrowid
    return {"id": new_id, "numero": numero, "mensaje": f"Período {numero:03d} creado"}

@app.put("/api/periodos_pago/{pp_id}")
def actualizar_periodo_pago(pp_id: int, data: PeriodoPagoUpdate):
    fields = {k: v for k, v in data.dict().items() if v is not None}
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    with get_db() as conn:
        cur = conn.cursor()
        if USE_PG:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            cur.execute(f"UPDATE periodos_pago SET {set_clause} WHERE id=%s", list(fields.values()) + [pp_id])
        else:
            set_clause = ", ".join(f"{k}=?" for k in fields)
            cur.execute(f"UPDATE periodos_pago SET {set_clause} WHERE id=?", list(fields.values()) + [pp_id])
    return {"mensaje": "Período actualizado"}

@app.post("/api/periodos_pago/{pp_id}/cerrar")
def cerrar_periodo_pago(pp_id: int, usuario: Optional[str] = None):
    with get_db() as conn:
        cur = conn.cursor()
        # Verificar pendientes
        cur.execute(f"SELECT COUNT(*) as n FROM tickets WHERE periodo_pago_id={PH} AND estado='pendiente'", (pp_id,))
        row = _row(cur)
        pendientes = row["n"] if row else 0
        if pendientes > 0:
            raise HTTPException(409, f"Hay {pendientes} ticket(s) pendientes en este período. Revisalos antes de cerrar.")
        cur.execute(f"UPDATE periodos_pago SET estado='cerrado', cerrado_por={PH}, cerrado_en=CURRENT_TIMESTAMP WHERE id={PH}",
                    (usuario, pp_id))
    return {"mensaje": "Período cerrado correctamente"}

@app.post("/api/periodos_pago/{pp_id}/reabrir")
def reabrir_periodo_pago(pp_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"UPDATE periodos_pago SET estado='abierto', cerrado_en=NULL, cerrado_por=NULL WHERE id={PH}", (pp_id,))
    return {"mensaje": "Período reabierto"}

@app.get("/api/periodos_pago/{pp_id}/resumen")
def resumen_periodo_pago(pp_id: int):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT a.id, a.nombre, a.tope_mensual, a.cbu, a.banco, a.alias, a.cuit,
                COALESCE(SUM(CASE WHEN t.estado='pendiente'      THEN t.valor          END),0) as total_pendiente,
                COALESCE(SUM(CASE WHEN t.estado='aprobado'       THEN t.valor_aprobado END),0) as total_aprobado,
                COALESCE(SUM(CASE WHEN t.estado='debito_parcial' THEN t.valor_aprobado END),0) as total_debito,
                COALESCE(SUM(CASE WHEN t.estado='pagado'         THEN t.valor_aprobado END),0) as total_pagado,
                COALESCE(SUM(CASE WHEN t.estado='rechazado'      THEN t.valor          END),0) as total_rechazado,
                COUNT(CASE WHEN t.estado='pendiente'      THEN 1 END) as cant_pendientes,
                COUNT(CASE WHEN t.estado='aprobado'       THEN 1 END) as cant_aprobados,
                COUNT(CASE WHEN t.estado='debito_parcial' THEN 1 END) as cant_debito,
                COUNT(CASE WHEN t.estado='pagado'         THEN 1 END) as cant_pagados,
                COUNT(CASE WHEN t.estado='rechazado'      THEN 1 END) as cant_rechazados
            FROM agentes a
            LEFT JOIN tickets t ON a.id = t.agente_id AND t.periodo_pago_id = {PH}
            WHERE a.activo = 1
            GROUP BY a.id, a.nombre, a.tope_mensual, a.cbu, a.banco, a.alias, a.cuit
            ORDER BY a.nombre
        """, (pp_id,))
        rows = _rows(cur)
        result = []
        for d in rows:
            subtotal = d["total_aprobado"] + d["total_debito"] + d["total_pagado"]
            d["tope_efectivo"] = d["tope_mensual"]
            d["a_transferir"] = min(subtotal, d["tope_mensual"]) if d["tope_mensual"] > 0 else subtotal
            d["excedente"] = max(0, subtotal - d["tope_mensual"]) if d["tope_mensual"] > 0 else 0
            result.append(d)
        return result

@app.get("/api/periodos_pago/{pp_id}/exportar/excel")
def exportar_excel_periodo_pago(pp_id: int):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment as XAlign
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM periodos_pago WHERE id={PH}", (pp_id,))
        pp = _row(cur)
        if not pp:
            raise HTTPException(404, "Período no encontrado")
    nombre_pp = f"{pp['numero']:03d} - {pp['nombre']}"
    resumen_data = resumen_periodo_pago(pp_id)
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT t.*, a.nombre as agente_nombre, c.nombre as categoria_nombre
            FROM tickets t
            LEFT JOIN agentes a ON t.agente_id = a.id
            LEFT JOIN categorias c ON t.categoria_id = c.id
            WHERE t.periodo_pago_id = {PH}
            ORDER BY a.nombre, t.fecha_gasto
        """, (pp_id,))
        tickets = _rows(cur)
        for t in tickets:
            if t.get("fecha_gasto") and not isinstance(t["fecha_gasto"], str):
                t["fecha_gasto"] = t["fecha_gasto"].strftime("%Y-%m-%d")

    wb = openpyxl.Workbook()

    # Hoja 1: Resumen por agente
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"SINDICATO ATE — PERÍODO {nombre_pp}"
    ws1["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws1["A1"].alignment = XAlign(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28
    hdrs1 = ["Agente", "CUIT", "CBU", "Banco/Alias", "Aprobado", "Pagado", "Rechazado", "A Transferir"]
    for c, h in enumerate(hdrs1, 1):
        cell = ws1.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2d6a9f")
        cell.alignment = XAlign(horizontal="center")
    total_transf = 0
    for ri, a in enumerate(resumen_data, 3):
        aprobado = a["total_aprobado"] + a["total_debito"]
        vals = [a["nombre"], a.get("cuit") or "-", a.get("cbu") or "-",
                a.get("banco") or a.get("alias") or "-",
                aprobado, a["total_pagado"], a["total_rechazado"], a["a_transferir"]]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=c, value=v)
            if c in (5, 6, 7, 8):
                cell.number_format = '"$"#,##0.00'
        total_transf += a["a_transferir"]
    tr = len(resumen_data) + 3
    ws1.cell(row=tr, column=7, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=tr, column=8, value=total_transf).number_format = '"$"#,##0.00'
    ws1.cell(row=tr, column=8).font = Font(bold=True)
    for col, w in zip("ABCDEFGH", [28, 16, 22, 18, 14, 14, 14, 14]):
        ws1.column_dimensions[col].width = w

    # Hoja 2: Detalle de tickets
    ws2 = wb.create_sheet("Detalle")
    ws2.merge_cells("A1:I1")
    ws2["A1"] = f"DETALLE DE TICKETS — PERÍODO {nombre_pp}"
    ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1a3a5c")
    ws2["A1"].alignment = XAlign(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26
    hdrs2 = ["Agente", "Fecha", "Categoría", "Comprobante", "Descripción", "Valor", "Estado", "Aprobado", "Motivo"]
    for c, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2d6a9f")
    colores = {"aprobado": "D4EDDA", "rechazado": "F8D7DA", "debito_parcial": "FFF3CD",
               "pendiente": "FFFFFF", "pagado": "E0F2FE"}
    for ri, t in enumerate(tickets, 3):
        color = colores.get(t["estado"], "FFFFFF")
        for c, v in enumerate([
            t["agente_nombre"], t["fecha_gasto"], t.get("categoria_nombre", ""),
            t.get("comprobante", ""), t.get("descripcion", ""), t["valor"],
            t["estado"].upper(), t.get("valor_aprobado") or "", t.get("motivo_rechazo", "")
        ], 1):
            cell = ws2.cell(row=ri, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=color)
            if c in (6, 8):
                cell.number_format = '"$"#,##0.00'
    for col, w in zip("ABCDEFGHI", [28, 12, 16, 18, 32, 14, 14, 14, 30]):
        ws2.column_dimensions[col].width = w

    filename = f"periodo_{pp['numero']:03d}_ATE.xlsx"
    path = os.path.join(EXPORTS_DIR, filename)
    wb.save(path)
    return FileResponse(path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
